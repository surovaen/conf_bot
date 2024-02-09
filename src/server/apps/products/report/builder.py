from pathlib import Path
from tempfile import NamedTemporaryFile
from typing import List

import openpyxl
from openpyxl.cell import Cell
from openpyxl.styles import Border, Font, Side


class ReportXlsxBuilder:
    """Класс построения отчета."""

    REPORT_TEMPLATE = Path(__file__).resolve().parent.joinpath('template.xlsx')

    TITLE_TEXT_CELL = 'B1'
    DATE_TEXT_CELL = 'B2'
    TIME_TEXT_CELL = 'B3'

    COLUMNS = range(1, 5)

    FONT = Font(name='Times New Roman', sz=14)
    BORDER = Side(border_style='thin', color='FF000000')

    def __init__(self, data: List[dict]):
        self._data = data[0]
        self.wb = openpyxl.open(self.REPORT_TEMPLATE, data_only=True)
        self.ws = self.wb.active

    def build(self) -> bytes:
        """Построение отчета."""

        self.ws[self.TITLE_TEXT_CELL] = self._data.get('title')
        self.ws[self.DATE_TEXT_CELL] = self._data.get('date')
        self.ws[self.TIME_TEXT_CELL] = self._data.get('time')

        for num, user in enumerate(self._data.get('users'), start=1):
            self._flush_row(num, user)

        with NamedTemporaryFile() as tmp:
            self.wb.save(tmp.name)
            tmp.seek(0)
            stream = tmp.read()

        return stream

    def _flush_row(self, num, user):
        """Метод заполнения и форматирования строк отчета."""
        row = [
            str(num),
            user['name'],
            user['phone_number'],
            user['menu'],
        ]

        self.ws.append(row)

        for col in self.COLUMNS:
            cell = self.ws.cell(row=self.ws._current_row, column=col)
            cell.font = self.FONT
            self._draw_border(cell)

    def _draw_border(self, cell: Cell):
        """Метод добавления границы вокруг ячейки."""
        cell.border = Border(
            top=self.BORDER,
            right=self.BORDER,
            bottom=self.BORDER,
            left=self.BORDER,
        )
